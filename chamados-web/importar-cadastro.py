# -*- coding: utf-8 -*-
"""Importa o cadastro de placas e de motoristas para o arquivo
chamados-data/cadastro.json, usado pelo autocompletar do formulário
"Novo chamado" e pela aba Telefones.

Fontes:
  1. FROTA CERTADOC (xlsx) — placas com renavam.
  2. PROGRAMAÇÃO BRAZIL TRANSPORTS (xlsx):
     - aba VEÍCULOS/CARRETAS — demais placas (cavalos e carretas);
     - aba MOTORISTAS — condutores (nome, CPF, telefone, vínculo, CNH);
     - aba CONTATOS — telefone de quem está sem na MOTORISTAS.

Uso (com uv, sem instalar nada):
    uv run --with openpyxl python importar-cadastro.py
    uv run --with openpyxl python importar-cadastro.py --frota "C:\\caminho\\FROTA.xlsx" --programacao "C:\\caminho\\PROGRAMACAO.xlsx"

Ou dê dois cliques em IMPORTAR-CADASTRO.bat.
Rode de novo sempre que a frota ou a programação mudar.
"""
import argparse
import json
import os
import re
import sys
import unicodedata
import warnings
from datetime import datetime

PADRAO_FROTA = r"C:\Users\brazil\Downloads\FROTA CERTADOC (2).xlsx"
PADRAO_PROGRAMACAO = r"C:\Users\brazil\Downloads\PROGRAMAÇÃO BRAZIL TRANSPORTS.xlsx"
RE_PLACA = re.compile(r"[A-Z]{3}\d[A-Z0-9]\d{2}")  # antiga (AAA0000) ou Mercosul (AAA0A00)
RAIZ = os.path.dirname(os.path.abspath(__file__))
PADRAO_SAIDA = os.path.join(os.path.dirname(RAIZ), "chamados-data", "cadastro.json")


def so_digitos(v):
    return re.sub(r"\D", "", v or "")


def fmt_telefone(v):
    d = so_digitos(v)
    if len(d) == 11:
        return f"({d[0:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 10:
        return f"({d[0:2]}) {d[2:6]}-{d[6:]}"
    return ""  # incompleto: melhor vazio do que travar o formulário


def fmt_cpf(v):
    d = so_digitos(v)
    if len(d) != 11:
        return ""
    return f"{d[0:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"


def ler_frota(caminho):
    import openpyxl
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    # A aba de dados é a "1º PASSO" (a primeira é só instruções).
    ws = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active
    veiculos = []
    vistos = set()
    for row in ws.iter_rows(min_row=10, values_only=True):
        placa = str(row[1] or "").strip().upper().replace("-", "")
        if not RE_PLACA.fullmatch(placa):
            continue  # ignora cabeçalho, vazios e lixo
        if placa in vistos:
            continue
        vistos.add(placa)
        veiculos.append({
            "placa": placa,
            "renavam": so_digitos(str(row[2] or "")),
        })
    return veiculos


def sem_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if not unicodedata.combining(c))


def ler_programacao(caminho):
    """Placas (cavalos e carretas) da planilha de programação.

    Procura primeiro a aba de veículos/carretas; se não existir, cai para a
    de motoristas. Varre todas as células da aba: qualquer valor no formato
    de placa entra, então a importação sobrevive a mudanças de layout.
    """
    import openpyxl
    warnings.filterwarnings("ignore", module="openpyxl")  # datas inválidas em outras abas
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    def acha_aba(*trechos):
        for nome in wb.sheetnames:
            chave = sem_acentos(nome).upper()
            if all(t in chave for t in trechos):
                return wb[nome]
        return None

    candidatas = [acha_aba("VEICULO"), acha_aba("CARRETA"), acha_aba("MOTORISTA")]
    placas = []
    vistos = set()
    for ws in candidatas:
        if ws is None:
            continue
        for row in ws.iter_rows(values_only=True):
            for cel in row:
                if cel is None:
                    continue
                v = str(cel).strip().upper().replace("-", "")
                if RE_PLACA.fullmatch(v) and v not in vistos:
                    vistos.add(v)
                    placas.append(v)
        if placas:
            break  # achou na primeira aba candidata: não mistura com as demais
    wb.close()
    return placas


def cel_texto(v):
    """Célula como texto. Números vêm do Excel como float (81994620905.0)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if hasattr(v, "strftime"):  # datas
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def ler_motoristas(caminho):
    """Condutores da aba MOTORISTAS da planilha de programação.

    Telefone: usa o da própria aba; quem está sem ganha o da aba CONTATOS
    (busca pelo nome). Duplicados (mesmo CPF ou mesmo nome) são unificados
    mantendo o registro mais completo.
    """
    import openpyxl
    warnings.filterwarnings("ignore", module="openpyxl")
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    def acha_aba(trecho):
        for nome in wb.sheetnames:
            if trecho in sem_acentos(nome).upper():
                return wb[nome]
        return None

    def nome_normal(v):
        return re.sub(r"\s+", " ", sem_acentos(v).strip().upper())

    # telefones da aba CONTATOS, indexados pelo nome
    contatos = {}
    aba_contatos = acha_aba("CONTATO")
    if aba_contatos is not None:
        col = None
        for row in aba_contatos.iter_rows(values_only=True):
            rotulos = [sem_acentos(cel_texto(c)).upper() for c in row]
            if col is None:
                if "NOME" in rotulos and "TELEFONE" in rotulos:
                    col = {"nome": rotulos.index("NOME"), "tel": rotulos.index("TELEFONE")}
                continue
            nome = nome_normal(cel_texto(row[col["nome"]]))
            tel = fmt_telefone(cel_texto(row[col["tel"]]))
            if nome and tel and nome not in contatos:
                contatos[nome] = tel

    aba = acha_aba("MOTORISTA")
    if aba is None:
        sys.exit("ERRO: aba MOTORISTAS não encontrada na planilha da programação")

    col = None
    motoristas = []
    por_chave = {}
    for row in aba.iter_rows(values_only=True):
        rotulos = [sem_acentos(cel_texto(c)).upper() for c in row]
        if col is None:
            # a linha de cabeçalho é a que tem a coluna MOTORISTA
            if "MOTORISTA" in rotulos:
                col = {"nome": rotulos.index("MOTORISTA")}
                for chave, trecho in [("status", "STATUS"), ("vinculo", "AGREGADO"),
                                      ("cpf", "CPF"), ("tel", "TELEFONE"), ("cnh", "VENCIMENTO CNH")]:
                    col[chave] = next((i for i, r in enumerate(rotulos) if trecho in r), None)
            continue

        def valor(chave):
            i = col.get(chave)
            return cel_texto(row[i]) if i is not None and i < len(row) else ""

        nome = re.sub(r"\s+", " ", valor("nome")).strip()
        if not nome:
            continue
        cpf_bruto = valor("cpf")
        if cpf_bruto and "." not in cpf_bruto and "-" not in cpf_bruto:
            # CPF gravado como número perde os zeros à esquerda
            cpf_bruto = so_digitos(cpf_bruto).zfill(11)
        m = {
            "nome": nome,
            "cpf": fmt_cpf(cpf_bruto),
            "telefone": fmt_telefone(valor("tel")) or contatos.get(nome_normal(nome), ""),
            "status": valor("status").upper(),
            "vinculo": valor("vinculo").upper(),
            "vencimentoCnh": valor("cnh"),
        }
        chave = so_digitos(m["cpf"]) or nome.upper()
        existente = por_chave.get(chave)
        if existente:
            # duplicado: completa o que faltar no registro já visto
            for campo in ("cpf", "telefone", "status", "vinculo", "vencimentoCnh"):
                if not existente[campo] and m[campo]:
                    existente[campo] = m[campo]
            continue
        por_chave[chave] = m
        motoristas.append(m)
    wb.close()
    motoristas.sort(key=lambda m: m["nome"].upper())
    return motoristas


def main():
    ap = argparse.ArgumentParser(description="Importa placas e motoristas para o cadastro.json")
    ap.add_argument("--frota", default=PADRAO_FROTA, help="planilha xlsx da frota (CertaDoc)")
    ap.add_argument("--programacao", default=PADRAO_PROGRAMACAO,
                    help="planilha xlsx da programação (abas VEÍCULOS/CARRETAS, MOTORISTAS e CONTATOS)")
    ap.add_argument("--saida", default=PADRAO_SAIDA, help="cadastro.json de destino")
    args = ap.parse_args()

    for caminho, rotulo in [(args.frota, "frota"), (args.programacao, "programação")]:
        if not os.path.isfile(caminho):
            sys.exit(f"ERRO: arquivo de {rotulo} não encontrado: {caminho}")

    veiculos = ler_frota(args.frota)  # CertaDoc primeiro: é quem tem o renavam
    ja_tem = {v["placa"] for v in veiculos}
    novas = [p for p in ler_programacao(args.programacao) if p not in ja_tem]
    veiculos += [{"placa": p, "renavam": ""} for p in novas]
    veiculos.sort(key=lambda v: v["placa"])
    motoristas = ler_motoristas(args.programacao)
    cadastro = {
        "atualizadoEm": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "fonteFrota": os.path.basename(args.frota),
        "fonteMotoristas": os.path.basename(args.programacao),
        "fonteProgramacao": os.path.basename(args.programacao),
        "veiculos": veiculos,
        "motoristas": motoristas,
    }
    os.makedirs(os.path.dirname(args.saida), exist_ok=True)
    tmp = args.saida + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cadastro, f, ensure_ascii=False, indent=1)
    os.replace(tmp, args.saida)
    print(f"OK: {len(veiculos)} veículos e {len(motoristas)} motoristas gravados em {args.saida}")


if __name__ == "__main__":
    main()
